#!/bin/python3
import yaml
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment


def read_yaml_file(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        result = yaml.load(f.read(), Loader=yaml.FullLoader)
    return result


def convert_sigs_dict_data_to_pkg_sig_dict(sigs_dict_data):
    sig_info_dict = {}
    for sig in sigs_dict_data["sigs"]:
        if sig["name"] in ["sig-UKUI", "sig-Ha", "sig-recycle", "sig-wine", "sig-aarch32"]:
            continue
        for pkg in sig["repositories"]:
            if pkg.startswith("src-openeuler") and not pkg.startswith("src-openeuler/dde") and not pkg.startswith(
                    "src-openeuler/deepin"):
                sig_info_dict[pkg.split('/')[1]] = sig["name"]
    return sig_info_dict


def dict_to_excel(data_dict, excel_name):
    if os.path.exists(excel_name):
        os.rename(excel_name, "{0}-{1}".format(datetime.now().strftime('%Y_%m_%d_%H_%M_%S'), excel_name))
        # os.remove(excel_name)
    wb = Workbook()
    ws = wb.active

    row = 2
    for pkg, sig in data_dict.items():
        ws.cell(row=row, column=1, value=pkg)
        ws.cell(row=row, column=2, value=sig)
        row += 1
    wb.save(excel_name)


# if __name__ == "__main__":
#     sigs_dict_data = read_yaml_file("sigs.yaml")
#     pkg_sig_dict = convert_sigs_dict_data_to_pkg_sig_dict(sigs_dict_data)
#     dict_to_excel(pkg_sig_dict, "test.xlsx")
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
)

wb = Workbook()
ws = wb.active

rows = [
    (None, 2013, 2014),
    ("Apples", 5, 4),
    ("Oranges", 6, 2),
    ("Pears", 8, 3)
]

for row in rows:
    ws.append(row)

data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=4)
titles = Reference(ws, min_col=1, min_row=2, max_row=4)
chart = BarChart3D()
chart.title = "3D Bar Chart"
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)

ws.add_chart(chart, "E5")
wb.save("bar3d.xlsx")